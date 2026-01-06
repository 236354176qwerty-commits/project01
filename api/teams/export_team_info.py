from flask import jsonify, session, send_file
from io import BytesIO
from datetime import datetime
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from database import DatabaseManager
from utils.decorators import log_action, handle_db_errors

from . import teams_bp


def _normalize_gender(value):
    if value is None:
        return ''
    s = str(value).strip()
    if not s:
        return ''
    lower = s.lower()
    if lower in ['male', 'm', '男']:
        return '男'
    if lower in ['female', 'f', '女']:
        return '女'
    return s


def _safe_filename(value: str) -> str:
    if not value:
        return '队伍'
    cleaned = ''.join(ch for ch in value if ch not in '\\/:*?"<>|')
    return cleaned.strip() or '队伍'


def _format_selected_events(selected_events, competition_event) -> str:
    def _join(items):
        cleaned = []
        for it in items:
            if it is None:
                continue
            s = str(it).strip()
            if not s:
                continue
            if s in ['和', '以及', '及', '&', 'and', 'AND', 'And']:
                continue
            cleaned.append(s)
        return '、'.join(cleaned)

    if isinstance(selected_events, (list, tuple, set)):
        joined = _join(list(selected_events))
        return joined or (competition_event or '')

    if isinstance(selected_events, str):
        s = selected_events.strip()
        if not s:
            return competition_event or ''

        parsed = None
        if s.startswith('[') and s.endswith(']'):
            try:
                parsed = json.loads(s)
            except Exception:
                try:
                    parsed = json.loads(s.replace("'", '"'))
                except Exception:
                    parsed = None

        if isinstance(parsed, list):
            joined = _join(parsed)
            return joined or (competition_event or '')

        return s

    return competition_event or ''


@teams_bp.route('/team/<int:team_id>/export', methods=['GET'])
@log_action('导出队伍信息')
@handle_db_errors
def api_export_team_info(team_id):
    """导出队伍信息表（xlsx）"""
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': '请先登录'}), 401

    user_role = session.get('user_role')
    current_user_id = session.get('user_id')

    db_manager = DatabaseManager()
    with db_manager.get_connection() as conn:
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT
                team_id,
                event_id,
                team_name,
                leader_name,
                leader_phone,
                leader_email,
                team_address,
                team_description,
                submitted_for_review,
                submitted_at,
                created_by
            FROM teams
            WHERE team_id = %s
            LIMIT 1
            """,
            (team_id,),
        )
        team = cursor.fetchone()
        if not team:
            cursor.close()
            return jsonify({'success': False, 'message': '队伍不存在'}), 404

        is_admin = user_role in ['admin', 'super_admin']
        if not is_admin and team.get('created_by') != current_user_id:
            cursor.close()
            return jsonify({'success': False, 'message': '您没有权限导出此队伍信息'}), 403

        event_id = team.get('event_id')

        cursor.execute(
            """
            SELECT
                name,
                gender,
                age,
                phone,
                id_card,
                competition_event,
                selected_events
            FROM team_players
            WHERE team_id = %s AND event_id = %s
            ORDER BY player_id ASC
            """,
            (team_id, event_id),
        )
        players = cursor.fetchall() or []

        cursor.execute(
            """
            SELECT
                name,
                position,
                gender,
                age,
                phone,
                id_card
            FROM team_staff
            WHERE team_id = %s AND event_id = %s AND status = 'active'
            ORDER BY staff_id ASC
            """,
            (team_id, event_id),
        )
        staff = cursor.fetchall() or []

        cursor.close()

    wb = Workbook()
    ws = wb.active
    ws.title = '队伍信息表'

    header_fill = PatternFill('solid', fgColor='0D6EFD')
    header_font = Font(color='FFFFFF', bold=True)
    title_font = Font(size=14, bold=True)
    bold_font = Font(bold=True)
    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cell(row, col, value, *, font=None, align=None, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.border = border
        if font:
            c.font = font
        if align:
            c.alignment = align
        if fill:
            c.fill = fill
        return c

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.merge_cells('A1:F1')
    set_cell(1, 1, '队伍信息表', font=title_font, align=center)

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    set_cell(row, 1, '队伍基本信息', font=header_font, fill=header_fill, align=left)
    row += 1

    submitted_at = team.get('submitted_at')
    if isinstance(submitted_at, datetime):
        submitted_at_str = submitted_at.strftime('%Y-%m-%d %H:%M')
    else:
        submitted_at_str = ''

    submit_status = '已提交' if team.get('submitted_for_review') else '未提交'

    # 前两行：三列排版（每列 2 个单元格：label/value）
    # 第一行：队伍名称、领队姓名、提交状态
    set_cell(row, 1, '队伍名称', font=bold_font, align=center)
    set_cell(row, 2, team.get('team_name') or '', align=left)
    set_cell(row, 3, '领队姓名', font=bold_font, align=center)
    set_cell(row, 4, team.get('leader_name') or '', align=left)
    set_cell(row, 5, '提交状态', font=bold_font, align=center)
    set_cell(row, 6, submit_status, align=center)
    row += 1

    # 第二行：联系电话、邮箱地址、提交时间
    set_cell(row, 1, '联系电话', font=bold_font, align=center)
    set_cell(row, 2, team.get('leader_phone') or '', align=left)
    set_cell(row, 3, '邮箱地址', font=bold_font, align=center)
    set_cell(row, 4, team.get('leader_email') or '', align=left)
    set_cell(row, 5, '提交时间', font=bold_font, align=center)
    set_cell(row, 6, submitted_at_str, align=center)
    row += 1

    # 地址/简介：单独两行，value 跨列
    set_cell(row, 1, '队伍地址', font=bold_font, align=center)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    set_cell(row, 2, team.get('team_address') or '', align=left)
    row += 1

    set_cell(row, 1, '队伍简介', font=bold_font, align=center)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    set_cell(row, 2, team.get('team_description') or '', align=left)
    row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    set_cell(row, 1, '参赛选手列表', font=header_font, fill=header_fill, align=left)
    row += 1

    player_headers = ['姓名', '身份证号', '联系电话', '性别', '年龄', '参赛项目']
    for idx, h in enumerate(player_headers, start=1):
        set_cell(row, idx, h, font=bold_font, align=center)
    row += 1

    if players:
        for p in players:
            events_text = _format_selected_events(
                p.get('selected_events'),
                p.get('competition_event'),
            )

            values = [
                p.get('name') or '',
                (p.get('id_card') or ''),
                p.get('phone') or '',
                _normalize_gender(p.get('gender')),
                p.get('age') if p.get('age') is not None else '',
                events_text,
            ]
            for idx, val in enumerate(values, start=1):
                set_cell(row, idx, val, align=left if idx in [1, 2, 6] else center)
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        set_cell(row, 1, '（暂无选手数据）', align=center)
        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    set_cell(row, 1, '随行人员列表', font=header_font, fill=header_fill, align=left)
    row += 1

    staff_headers = ['姓名', '身份证号', '联系电话', '性别', '年龄', '角色/职务']
    for idx, h in enumerate(staff_headers, start=1):
        set_cell(row, idx, h, font=bold_font, align=center)
    row += 1

    if staff:
        position_map = {
            'coach': '教练',
            'head_coach': '教练',
            'manager': '领队',
            'medical': '医务人员',
            'doctor': '医务人员',
            'staff': '随行人员',
        }
        for s in staff:
            pos = s.get('position') or ''
            pos_label = position_map.get(str(pos).lower(), pos)
            values = [
                s.get('name') or '',
                s.get('id_card') or '',
                s.get('phone') or '',
                _normalize_gender(s.get('gender')),
                s.get('age') if s.get('age') is not None else '',
                pos_label,
            ]
            for idx, val in enumerate(values, start=1):
                set_cell(row, idx, val, align=left if idx in [1, 2, 6] else center)
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        set_cell(row, 1, '（暂无随行人员数据）', align=center)
        row += 1

    # Column widths
    widths = [14, 20, 14, 22, 14, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w

    # Row heights for header blocks
    ws.row_dimensions[1].height = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{_safe_filename(team.get('team_name'))}_队伍信息表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
