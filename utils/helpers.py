#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
武术赛事管理系统 - 辅助函数
"""

import os
import uuid
import hashlib
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from flask import current_app
import re

def generate_unique_filename(filename):
    """生成唯一的文件名"""
    if filename:
        # 获取文件扩展名
        ext = os.path.splitext(filename)[1]
        # 生成唯一标识符
        unique_id = str(uuid.uuid4())
        # 生成时间戳
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{timestamp}_{unique_id}{ext}"
    return None

def allowed_file(filename):
    """检查文件类型是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

def save_uploaded_file(file, subfolder=''):
    """保存上传的文件"""
    if file and allowed_file(file.filename):
        # 生成安全的文件名
        original_filename = secure_filename(file.filename)
        unique_filename = generate_unique_filename(original_filename)
        
        # 创建保存路径
        upload_folder = current_app.config['UPLOAD_FOLDER']
        if subfolder:
            upload_folder = os.path.join(upload_folder, subfolder)
            os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        
        # 保存文件
        file.save(file_path)
        
        return {
            'success': True,
            'filename': unique_filename,
            'original_filename': original_filename,
            'file_path': file_path,
            'relative_path': os.path.join(subfolder, unique_filename) if subfolder else unique_filename
        }
    
    return {'success': False, 'error': '文件类型不支持'}

def generate_registration_number(event_id, participant_count):
    """生成参赛编号"""
    # 格式: E{event_id:03d}P{participant_count:04d}
    return f"E{event_id:03d}P{participant_count:04d}"

def calculate_age(birth_date):
    """计算年龄"""
    if not birth_date:
        return None
    
    today = datetime.now().date()
    if isinstance(birth_date, datetime):
        birth_date = birth_date.date()
    
    age = today.year - birth_date.year
    if today.month < birth_date.month or (today.month == birth_date.month and today.day < birth_date.day):
        age -= 1
    
    return age

def get_age_group(birth_date):
    """根据出生日期获取年龄组"""
    age = calculate_age(birth_date)
    if not age:
        return None
    
    if age <= 12:
        return '儿童组(6-12岁)'
    elif age <= 17:
        return '少年组(13-17岁)'
    elif age <= 35:
        return '青年组(18-35岁)'
    elif age <= 50:
        return '中年组(36-50岁)'
    else:
        return '老年组(51岁以上)'

def format_datetime(dt, format_str='%Y-%m-%d %H:%M:%S'):
    """格式化日期时间"""
    if not dt:
        return ''
    
    if isinstance(dt, str):
        return dt
    
    return dt.strftime(format_str)

def format_date(date, format_str='%Y-%m-%d'):
    """格式化日期"""
    if not date:
        return ''
    
    if isinstance(date, str):
        return date
    
    if isinstance(date, datetime):
        date = date.date()
    
    return date.strftime(format_str)

def parse_datetime(date_str, format_str=None):
    """解析日期时间字符串（支持多种格式）"""
    if not date_str:
        return None
    
    # 如果指定了格式，直接使用
    if format_str:
        try:
            return datetime.strptime(date_str, format_str)
        except ValueError:
            return None
    
    # 尝试多种常见格式
    formats = [
        '%Y-%m-%dT%H:%M:%S',      # ISO格式: 2024-06-01T10:00:00
        '%Y-%m-%dT%H:%M:%S.%f',   # ISO格式带微秒: 2024-06-01T10:00:00.000
        '%Y-%m-%dT%H:%M',         # ISO格式无秒: 2024-06-01T10:00
        '%Y-%m-%d %H:%M:%S',      # 标准格式: 2024-06-01 10:00:00
        '%Y-%m-%d %H:%M',         # 无秒: 2024-06-01 10:00
        '%Y-%m-%d',               # 仅日期: 2024-06-01
    ]
    
    # 处理时区标识符
    date_str_clean = date_str.replace('Z', '').replace('+00:00', '')
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str_clean, fmt)
        except ValueError:
            continue
    
    # 尝试使用 fromisoformat（Python 3.7+）
    try:
        # 处理常见的ISO格式变体
        iso_str = date_str.replace('Z', '+00:00')
        return datetime.fromisoformat(iso_str.replace('+00:00', ''))
    except (ValueError, AttributeError):
        pass
    
    return None


def get_event_categories():
    """获取赛事分类选项
    返回传统项目、自选和规定项目、对练项目三个大类别的赛事分类
    """
    return [
        {'value': 'traditional', 'label': '传统项目'},
        {'value': 'optional_standard', 'label': '自选和规定项目'},
        {'value': 'dueling', 'label': '对练项目'}
    ]

def validate_email(email):
    """验证邮箱格式"""
    if not email:
        return False
    
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_phone(phone):
    """验证手机号格式"""
    if not phone:
        return False
    
    # 中国手机号格式
    pattern = r'^1[3-9]\d{9}$'
    return re.match(pattern, phone) is not None

def generate_password_hash(password, salt_length=16):
    """生成密码哈希

    返回 salt+hash 的十六进制字符串（仅包含 ASCII 字符），以避免在 utf8mb4 连接下
    向 MySQL 发送任意二进制数据导致 1300 Invalid utf8mb4 character string 错误。
    """
    # 生成随机盐
    salt = os.urandom(salt_length)
    # 使用 PBKDF2 算法生成哈希
    password_hash = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000)
    # 拼接盐和哈希后以十六进制字符串形式返回
    data = salt + password_hash
    return data.hex()


def verify_password(password, password_hash):
    """验证密码

    支持两种存储格式：
    1）新格式：salt+hash 的十六进制字符串（或其字节形式，例如从 VARBINARY 读出的 ASCII 字节）；
    2）旧格式：直接存储的原始二进制 salt+hash（长度约 48 字节）。
    """
    if not password_hash:
        return False

    raw = None

    # 从数据库读取时，VARBINARY 通常会得到 bytes/bytearray
    if isinstance(password_hash, (bytes, bytearray)):
        # 优先按十六进制 ASCII 解码（新格式）
        try:
            hex_str = password_hash.decode('ascii')
            raw = bytes.fromhex(hex_str)
        except (UnicodeDecodeError, ValueError):
            # 无法按十六进制解析时，退回为旧格式的原始二进制
            raw = bytes(password_hash)

    elif isinstance(password_hash, str):
        # 字符串：视为十六进制编码
        try:
            raw = bytes.fromhex(password_hash)
        except ValueError:
            return False
    else:
        # 其它类型暂不支持
        return False

    # 原始数据至少应包含 16 字节盐 + 32 字节哈希
    if not raw or len(raw) < 16 + 32:
        return False

    salt = raw[:16]
    stored_hash = raw[16:]
    computed_hash = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000)
    return computed_hash == stored_hash

def calculate_average_score(scores, drop_highest=True, drop_lowest=True):
    """计算平均分（可选择去掉最高分和最低分）"""
    if not scores:
        return 0.0
    
    # 转换为浮点数列表
    score_list = [float(score) for score in scores if score is not None]
    
    if len(score_list) == 0:
        return 0.0
    
    # 如果分数少于3个，不去掉最高最低分
    if len(score_list) < 3:
        return round(sum(score_list) / len(score_list), 2)
    
    # 排序
    score_list.sort()
    
    # 去掉最高分和最低分
    if drop_lowest:
        score_list = score_list[1:]
    if drop_highest and len(score_list) > 1:
        score_list = score_list[:-1]
    
    # 计算平均分
    if score_list:
        return round(sum(score_list) / len(score_list), 2)
    else:
        return 0.0

def format_score(score, decimal_places=2):
    """格式化分数显示"""
    if score is None:
        return '0.00'
    
    return f"{float(score):.{decimal_places}f}"

def get_ranking_suffix(rank):
    """获取排名后缀"""
    if rank == 1:
        return '🥇'
    elif rank == 2:
        return '🥈'
    elif rank == 3:
        return '🥉'
    else:
        return f"第{rank}名"

def paginate_list(items, page=1, per_page=20):
    """列表分页"""
    total = len(items)
    start = (page - 1) * per_page
    end = start + per_page
    
    return {
        'items': items[start:end],
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'has_prev': page > 1,
        'has_next': end < total,
        'prev_num': page - 1 if page > 1 else None,
        'next_num': page + 1 if end < total else None
    }

def export_to_excel(data, filename, sheet_name='Sheet1'):
    """导出数据到Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            return False
        
        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # 写入数据
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item.get(header, ''))
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(filename)
        return True
        
    except ImportError:
        return False
    except Exception as e:
        print(f"导出Excel失败: {e}")
        return False

def generate_qr_code(data, filename):
    """生成二维码"""
    try:
        import qrcode
        from PIL import Image
        
        # 创建二维码实例
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        
        # 添加数据
        qr.add_data(data)
        qr.make(fit=True)
        
        # 创建图像
        img = qr.make_image(fill_color="black", back_color="white")
        
        # 保存图像
        img.save(filename)
        return True
        
    except ImportError:
        return False
    except Exception as e:
        print(f"生成二维码失败: {e}")
        return False

def send_notification_email(to_email, subject, body, html_body=None):
    """发送通知邮件"""
    try:
        from flask_mail import Message, Mail
        
        mail = Mail(current_app)
        
        msg = Message(
            subject=subject,
            recipients=[to_email],
            body=body,
            html=html_body
        )
        
        mail.send(msg)
        return True
        
    except Exception as e:
        print(f"发送邮件失败: {e}")
        return False

def log_user_action(user_id, action, details=None):
    """记录用户操作日志"""
    try:
        import logging
        
        logger = logging.getLogger('user_actions')
        
        log_entry = {
            'user_id': user_id,
            'action': action,
            'timestamp': datetime.now().isoformat(),
            'details': details
        }
        
        logger.info(f"用户操作: {log_entry}")
        return True
        
    except Exception as e:
        print(f"记录日志失败: {e}")
        return False
